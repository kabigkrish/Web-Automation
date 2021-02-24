import time
import openpyxl
from dataclasses import asdict
#LOADING EXCEL AND CONVERTING TO DATACLASS
from datetime import datetime
from openpyxl import load_workbook
from classes import members
from mapping import plan_price,member_number,first_name,middle_name,last_name,date_of_birth,address_1,address_2,city,state,zip_code,email_address,contact_number, beneficiary_first_name,beneficiary_middle_name,beneficiary_last_name,relationship,percentage_of_benefit

workbook = load_workbook(filename="testing.xlsx", read_only=True)
sheet = workbook.active
member_class=[]
for row in sheet.iter_rows(min_row=2,min_col=1,max_row=3,values_only=True):
    member=members(plan_price=row[plan_price],
    member_number=row[member_number],
    first_name=row[first_name],
    middle_name=row[middle_name],
    last_name=row[last_name],
    date_of_birth=row[date_of_birth],
    address_1=row[address_1],
    address_2=row[address_2],
    city=row[city],
    state=row[state],
    zip_code=row[zip_code],
    email_address=row[email_address],
    contact_number=row[contact_number],
    beneficiary_last_name=row[beneficiary_last_name],
    beneficiary_first_name=row[beneficiary_first_name],
    beneficiary_middle_name=row[beneficiary_middle_name],
    relationship=row[relationship],
    percentage_of_benefit=row[percentage_of_benefit]
    )
    member_class.append(member)
print(len(member_class))

for i in range(0,2):
    a=asdict(member_class[i])
    print(a.get('beneficiary_first_name'))
